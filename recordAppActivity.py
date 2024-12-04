import os
import signal
import time
from threading import Thread, Lock
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import timedelta
import psutil
import win32gui
import win32process
import pyautogui

# Timestamp for the log file
#timestampFile = datetime.now().strftime("%H-%M_%d.%m")
timestampFile = datetime.now().strftime("%d.%m.%Y")
file_path = f"{timestampFile}.xlsx"
# Consider using a configuration file or environment variables
SLEEPING_TIME = 2
MAX_IDLE_TIME = 60
EXCLUDED_APPS = [
    'chrome',
    'vlc',
    'mbc-bex64',
    'mbc-be',
    'mpc-be64'
]
print(file_path)
print("Application is starting at " + str(datetime.now().strftime("%H:%M")))

# Initialize usage records and locks
usage_records = {}
usage_lock = Lock()
sleepingTime = 2

# Global flag to stop threads
stop_threads = False


def signal_handler(sig, frame):
    global stop_threads
    print("\n[INFO] Stopping application and finalizing data...")
    stop_threads = True

    # Wait for threads to stop
    tracking_thread.join(timeout=3)
    write_thread.join(timeout=3)

    # Aggregate data
    aggregate_detailed_usage()
    add_total_duration_to_sheet(file_path)
    # Open the log file after saving
    os.startfile(file_path)
    print("[INFO] Application exited.")
    exit(0)


def create_or_append_workbook():
    """Check if workbook exists, create if not, otherwise load existing workbook."""
    if os.path.exists(file_path):
        print("[INFO] Workbook exists, loading existing workbook...")
        wb = openpyxl.load_workbook(file_path)

        # Ensure Session_Log sheet exists
        if 'Session_Log' not in wb.sheetnames:
            ws = wb.create_sheet(title="Session_Log")
            ws.append(["App Name", "App Title", "Duration", "Timestamp"])
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Create Session_Log sheet
        ws = wb.create_sheet(title="Session_Log")
        ws.append(["App Name", "App Title", "Duration", "Timestamp"])

    wb.save(file_path)



def format_duration(seconds):
    """Format duration as HH:MM:SS."""
    hours, seconds = divmod(seconds, 3600)
    minutes, seconds = divmod(seconds, 60)
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"


def write_to_file():
    """Periodically write usage records to the log file."""
    global usage_records, stop_threads
    while not stop_threads:
        with usage_lock:
            if usage_records:
                try:
                    append_current_session_data()
                except Exception as e:
                    print(f"[Error] Writing to file: {e}")

            time.sleep(sleepingTime)

def append_current_session_data():
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Session_Log"]

    # Append current session data
    for (app_name, app_title), duration in usage_records.items():
        timestamp = datetime.now().strftime("%H:%M:%S %d-%m-%Y")
        ws.append([app_name, app_title, format_duration(duration), timestamp])

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)
    usage_records.clear()

def aggregate_detailed_usage():
    """Aggregate usage data into a detailed summary."""
    try:
        wb = openpyxl.load_workbook(file_path)
        app_consolidated_records = {}

        if "Session_Log" in wb.sheetnames:
            log_sheet = wb["Session_Log"]

            for row in log_sheet.iter_rows(min_row=2, values_only=True):
                app_name, app_title, duration, timestamp = row
                hours, minutes, seconds = map(int, duration.split(':'))
                duration_in_seconds = hours * 3600 + minutes * 60 + seconds

                key = (app_name, app_title)
                if key not in app_consolidated_records:
                    app_consolidated_records[key] = {'total_duration': duration_in_seconds, 'sessions': 1}
                else:
                    record = app_consolidated_records[key]
                    record['total_duration'] += duration_in_seconds
                    record['sessions'] += 1

        # 1. Check if the sheet exists
        if "Detailed_Summary" not in wb.sheetnames:
            detailed_sheet = wb.create_sheet(title="Detailed_Summary")
        else:
            detailed_sheet = wb["Detailed_Summary"]
            detailed_sheet.delete_rows(1, detailed_sheet.max_row)

        detailed_sheet.append(["App Name", "App Title", "Total Duration", "Sessions"])

        for (app_name, app_title), record in app_consolidated_records.items():
            detailed_sheet.append([app_name, app_title, format_duration(record['total_duration']), record['sessions']])

        # Auto-adjust column widths
        for col in detailed_sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            col_letter = get_column_letter(col[0].column)
            detailed_sheet.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_path)
        print("[INFO] Detailed usage summary created.")

    except Exception as e:
        print(f"[Error] Creating detailed summary: {e}")



def get_active_window():
    """Returns the active window's application name and title."""
    hwnd = win32gui.GetForegroundWindow()
    if hwnd and not win32gui.IsIconic(hwnd):  # Check if the window is not minimized
        _, process_id = win32process.GetWindowThreadProcessId(hwnd)
        process = psutil.Process(process_id)
        current_app = process.name()
        current_title = truncate_title(win32gui.GetWindowText(hwnd))
        return current_app, current_title
    elif win32gui.IsIconic(hwnd):
        print("Current window is minimized. Putting the thread to sleep.")
    return "Unknown", "No Active Window"


def is_excluded_app(current_app, current_title):
    #haven't checked if it works
    return any(
        excluded.lower() in current_app.lower() or
        excluded.lower() in current_title.lower()
        for excluded in EXCLUDED_APPS
    )


def track_app_usage():
    last_app = None
    last_title = None
    start_time = None
    idle_time = 0
    last_mouse_position = pyautogui.position()

    try:
        while not stop_threads:
            try:
                # Get active window and check for exclusions
                current_app, current_title = get_active_window()

                # Check if the active app is excluded (YouTube, Netflix, VLC, or MBC-be x64)
                if is_excluded_app(current_app, current_title):
                    idle_time = 0  # Reset idle time if excluded app is active

                current_time = time.time()

                # Initialize tracking for the first app
                if start_time is None:
                    last_app = current_app
                    last_title = current_title
                    start_time = current_time
                    continue

                # Detect app switch or long inactivity
                # todo fix wrong application being logged
                if current_app != last_app or current_title != last_title:
                    # Log the app usage duration before switching
                    duration = current_time - start_time
                    if duration > 0.0:
                        print(f"[INFO] App: {last_app}, Title: {last_title}, Duration: {format_duration(duration)}")
                        print("initiating logging")
                        log_usage(last_app, last_title, duration)

                    # Update the app details
                    last_app = current_app
                    last_title = current_title
                    start_time = current_time

                idle_time = handle_idle_time(last_mouse_position, idle_time, current_app, current_title)

                time.sleep(sleepingTime)

            except Exception as e:
                print(f"[Error] {e}")
                time.sleep(sleepingTime)
    finally:
        # Log final usage when thread stops
        if last_app and start_time:
            try:
                final_duration = time.time() - start_time
                if final_duration > 0:
                    print(f"[INFO] Final App: {last_app}, Final Title: {last_title}, Final Duration: {format_duration(final_duration)}")
                    log_usage(last_app, last_title, final_duration)
            except Exception as e:
                print(f"[Error] Logging final app usage: {e}")

            # Write the final records to the Excel file
            with usage_lock:
                try:
                    append_current_session_data()
                except Exception as e:
                    print(f"[Error] Writing final data to file: {e}")



def handle_idle_time(last_mouse_position, idle_time, current_app, current_title):
    # Track idle time
    is_idle, last_mouse_position = track_idle_time(last_mouse_position)
    if is_idle:
        idle_time += sleepingTime
    else:
        idle_time = 0  # Reset idle time if there's activity

    if idle_time >= MAX_IDLE_TIME:
        print("[INFO] PC is idle for more than 1 minute.")
        # Log the idle time if conditions allow
        if not is_excluded_app(current_app, current_title):
            log_usage("Idle", "Idle", idle_time)
        idle_time = 0  # Reset idle time after logging
        return idle_time
    else:
        return idle_time

def track_idle_time(last_mouse_position):
    """Tracks the idle time based on mouse position and resets if necessary."""
    current_mouse_position = pyautogui.position()
    if current_mouse_position == last_mouse_position:  # No mouse movement
        return True, current_mouse_position  # Increment idle time
    return False, current_mouse_position  # Reset idle time if there's activity


def log_usage(app_name, app_title, duration):
    """Log the usage of an application."""
    global usage_records
    key = (app_name, app_title)
    with usage_lock:
        if key in usage_records:
            usage_records[key] += duration
        else:
            usage_records[key] = duration



def truncate_title(title, max_length=50):
    """Truncate the window title to a specified maximum length."""
    if not title:
        return title

    if len(title) <= max_length:
        return title

    return title[:max_length-3] + "..."


# Function to convert time string to timedelta object
def time_to_timedelta(time_str):
    # Check if the time_str matches a valid time format (HH:MM:SS or MM:SS or SS)
    try:
        time_parts = list(map(int, time_str.split(':')))
        if len(time_parts) == 3:
            return timedelta(hours=time_parts[0], minutes=time_parts[1], seconds=time_parts[2])
        elif len(time_parts) == 2:
            return timedelta(minutes=time_parts[0], seconds=time_parts[1])
        elif len(time_parts) == 1:
            return timedelta(seconds=time_parts[0])
    except ValueError:
        # If it's not a valid time string, return zero timedelta
        print(f"Skipping invalid time format: {time_str}")
        return timedelta()


# Function to add the sum of durations to the second sheet
def add_total_duration_to_sheet(file_path):
    previous_total_time_passed = ""
    # Open the workbook and select the second sheet
    wb = openpyxl.load_workbook(file_path)

    # Use Session_Log sheet for calculations if Detailed_Summary exists
    if "Detailed_Summary" in wb.sheetnames:
        sheet = wb["Detailed_Summary"]

        # Remove any existing total duration rows
        rows_to_remove = []
        for row in range(sheet.max_row, 1, -1):
            if sheet.cell(row=row, column=1).value in ["Total Duration", last_Record]:
                rows_to_remove.append(row)
            if sheet.cell(row=row, column=1).value in [last_Record]:
                previous_total_time_passed = sheet.cell(row=row, column=2).value
                print("previous total time since passed since start is: " + previous_total_time_passed)


        # Remove rows from bottom to top to avoid index shifting
        for row in sorted(rows_to_remove, reverse=True):
            sheet.delete_rows(row)

        # Initialize total duration
        total_duration = timedelta()
        # Read durations from the sheet (assuming durations start from row 1 in column 2)
        for row in range(2, sheet.max_row + 1):
            duration_str = sheet.cell(row=row, column=3).value
            if duration_str:
                total_duration += time_to_timedelta(duration_str)
        # Convert total_duration back to string in the format HH:MM:SS
        total_duration_str = str(total_duration)

        # leave an empty space amd then Find the next empty row in the second sheet
        next_row = sheet.max_row + 2
        # Add the total duration to the last row
        sheet.cell(row=next_row, column=2, value="Total Duration")
        sheet.cell(row=next_row, column=3, value=total_duration_str)


        if not previous_total_time_passed:
            previous_total_time_passed = timedelta()

        # add total time passed from the start of the program to determine if the app tracking worked correctly (epalithefsi)
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=2, value=last_Record)
        total_time_passed = previous_total_time_passed + time_to_timedelta(str(int((time.time() - program_start_time))))
        sheet.cell(row=next_row, column=3, value=format_duration(int(total_time_passed.total_seconds()) ))

        # Save the workbook without the extra rows
        wb.save(file_path)

    print("[INFO] Removed extra duration rows.")


last_Record = "time passed in last execution"

program_start_time = time.time()

# Create the workbook and register signal handler
create_or_append_workbook()
signal.signal(signal.SIGINT, signal_handler)

# Start tracking and writing threads
tracking_thread = Thread(target=track_app_usage, daemon=True)
write_thread = Thread(target=write_to_file, daemon=True)
tracking_thread.start()
write_thread.start()

# Keep the main thread alive
try:
    while not stop_threads:
        time.sleep(SLEEPING_TIME)
except KeyboardInterrupt:
    signal_handler(None,None)